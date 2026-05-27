from pathlib import Path
import pandas as pd
from src.config import CONFIG

from openpyxl.styles import (
    Font,
    PatternFill,
    Alignment
)
from openpyxl.chart import PieChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.chart.label import DataLabelList


def normalize_product_name(product_name):
    """Normaliza nombres de productos para comparación consistente"""
    if pd.isna(product_name):
        return ""
    return (
        str(product_name)
        .strip()
        .lower()
        .replace("café", "cafe")
        .replace("  ", " ")
    )


def build_store_report(
    df,
):

    # Normalizar nombres de productos para evitar duplicados por diferencias de casing/espacios
    df_copy = df.copy()
    df_copy["producto_normalizado"] = df_copy[CONFIG.product_column].apply(
        normalize_product_name
    )

    # Agrupar por fecha, hora y producto normalizado, contando ocurrencias
    sales_counts = df_copy.groupby(
        [
            "Fecha",
            "Hora",
            "producto_normalizado",
        ],
        as_index=False,
    ).size().rename(columns={"size": "Cant Venta"})

    # Obtener el producto original (sin normalizar) para cada grupo
    product_map = (
        df_copy.groupby("producto_normalizado")[CONFIG.product_column]
        .first()
        .to_dict()
    )

    sales_counts["Producto"] = sales_counts["producto_normalizado"].map(
        product_map)

    # Mantener solo las columnas requeridas
    grouped = sales_counts[
        ["Fecha", "Hora", "Producto", "Cant Venta"]
    ].copy()

    rows = []

    dates = sorted(
        grouped[
            "Fecha"
        ]
        .unique()
    )

    for date in dates:

        day = (

            grouped[
                grouped[
                    "Fecha"
                ]
                ==
                date
            ]

            .copy()

            .sort_values(
                [
                    "Hora",
                    "Cant Venta",
                    "Producto",
                ],
                ascending=[
                    True,
                    False,
                    True,
                ],
            )
        )

        rows.append(
            day
        )

        interval_resume = (

            day.groupby(
                "Hora"
            )

            [
                "Cant Venta"
            ]

            .sum()

            .sort_values(
                ascending=False
            )
        )

        if (
            interval_resume.empty
        ):

            continue

        best_hour = (
            interval_resume
            .index[0]
        )

        rows.append(

            pd.DataFrame(
                [
                    {
                        "Producto":
                            "INTERVALO MÁS VENDIDO",

                        "Fecha":
                            "",

                        "Hora":
                            best_hour,
                    }
                ]
            )
        )

    return pd.concat(
        rows,
        ignore_index=True,
    )


def sanitize_filename(
    value,
):

    invalid = (
        '<>:"/\\|?*'
    )

    for char in invalid:

        value = (
            value.replace(
                char,
                "_",
            )
        )

    return value


def export_reports(
    df,
):

    folder = Path(
        CONFIG.output_folder
    )

    folder.mkdir(
        exist_ok=True
    )

    stores = (
        df[
            CONFIG.store_column
        ]
        .dropna()
        .unique()
    )

    for store in stores:

        print(
            f"Generando {store}"
        )

        # 1. Conservamos el reporte original para cálculos limpios
        raw_report = build_store_report(
            df[df[CONFIG.store_column] == store]
        )

        report = raw_report[
            [
                "Producto",
                "Fecha",
                "Hora",
                "Cant Venta",
            ]
        ].copy()

        path = (
            folder
            /
            (
                sanitize_filename(
                    str(
                        store
                    )
                )
                +
                ".xlsx"
            )
        )

        with pd.ExcelWriter(
            path,
            engine="openpyxl",
        ) as writer:

            report.to_excel(
                writer,
                index=False,
            )

            ws = (
                writer.sheets[
                    "Sheet1"
                ]
            )

            fill = (
                PatternFill(
                    start_color="FFF2CC",
                    end_color="FFF2CC",
                    fill_type="solid",
                )
            )

            font = (
                Font(
                    bold=True,
                )
            )

            align = (
                Alignment(
                    horizontal="center",
                    vertical="center",
                )
            )

            # --- NUEVA LÓGICA CORREGIDA PARA GRÁFICAS Y DISEÑO ---

            # Estilizado de filas de resumen (INTERVALO MÁS VENDIDO)
            start_row = 2
            current_date_str = ""

            # Estilizado de filas de resumen y detección de bloques
            for row in range(
                2,
                ws.max_row + 1,
            ):
                val_fecha = ws.cell(row, 2).value
                if val_fecha and val_fecha != "":
                    current_date_str = str(val_fecha)

                if (
                    ws.cell(
                        row,
                        1,
                    ).value
                    !=
                    "INTERVALO MÁS VENDIDO"
                ):
                    continue

                interval = (
                    ws.cell(
                        row,
                        3,
                    ).value
                )

                # limpiar fila
                for col in range(
                    1,
                    5,
                ):
                    ws.cell(
                        row,
                        col,
                    ).value = None

                # combinar
                ws.merge_cells(
                    f"A{row}:B{row}"
                )

                ws.merge_cells(
                    f"C{row}:D{row}"
                )

                ws[
                    f"A{row}"
                ] = (
                    "INTERVALO MÁS VENDIDO"
                )

                ws[
                    f"C{row}"
                ] = (
                    interval
                )

                for col in (
                    "A",
                    "C",
                ):

                    cell = (
                        ws[
                            f"{col}{row}"
                        ]
                    )

                    cell.fill = (
                        fill
                    )

                    cell.font = (
                        font
                    )

                    cell.alignment = (
                        align
                    )

                start_row = row + 1

            # --- GRÁFICOS SIMPLES Y DIRECTOS ---
            
            # Calcular totales por fecha/hora desde raw_report
            raw_report["Cant Venta"] = pd.to_numeric(
                raw_report["Cant Venta"], errors="coerce").fillna(0).astype(int)
            
            hourly_totals = raw_report.groupby(["Fecha", "Hora"])["Cant Venta"].sum().reset_index()
            
            # Obtener bloques de fechas (detectar dónde cambia la fecha en el Excel)
            date_blocks = {}
            current_date = None
            block_start = 2
            
            for row in range(2, ws.max_row + 1):
                cell_date = ws.cell(row, 2).value
                
                if cell_date is None or cell_date == "Fecha":
                    continue
                
                cell_date_str = str(cell_date).split()[0]  # Extraer solo la fecha, sin hora
                
                if current_date != cell_date_str:
                    if current_date is not None:
                        if current_date not in date_blocks:
                            date_blocks[current_date] = []
                        date_blocks[current_date].append((block_start, row - 1))
                    
                    current_date = cell_date_str
                    block_start = row
            
            # Agregar el último bloque
            if current_date is not None and current_date not in date_blocks:
                date_blocks[current_date] = []
            if current_date is not None:
                date_blocks[current_date].append((block_start, ws.max_row))
            
            # Generar gráficos para cada fecha
            chart_row = 2
            aux_col_row = 2
            
            for date_key in sorted(date_blocks.keys()):
                # Obtener datos de esta fecha
                date_data = hourly_totals[hourly_totals["Fecha"].astype(str).str.contains(date_key)]
                
                if date_data.empty:
                    continue
                
                # Escribir tabla auxiliar
                ws.cell(aux_col_row, 11).value = "Hora"
                ws.cell(aux_col_row, 12).value = "Ventas"
                aux_col_row += 1
                
                start_data_row = aux_col_row
                for _, row_data in date_data.iterrows():
                    ws.cell(aux_col_row, 11).value = row_data["Hora"]
                    ws.cell(aux_col_row, 12).value = int(row_data["Cant Venta"])
                    aux_col_row += 1
                
                end_data_row = aux_col_row - 1
                aux_col_row += 2  # Espacio entre tablas
                
                # Crear gráfico
                chart = PieChart()
                chart.title = f"Ventas por Hora - {date_key}"
                chart.height = 12
                chart.width = 18
                
                # Referencias a los datos
                data_ref = Reference(ws, min_col=12, min_row=start_data_row - 1, max_row=end_data_row)
                cats_ref = Reference(ws, min_col=11, min_row=start_data_row, max_row=end_data_row)
                
                chart.add_data(data_ref, titles_from_data=True)
                chart.set_categories(cats_ref)
                
                # Configurar etiquetas
                chart.dataLabels = DataLabelList()
                chart.dataLabels.showCatName = False
                chart.dataLabels.showVal = True
                
                ws.add_chart(chart, f"F{chart_row}")
                chart_row += 25

            # Ajuste de ancho de columnas automático
            for column in ws.columns:
                try:
                    width = max(
                        len(str(cell.value or ""))
                        for cell in column
                        if cell.coordinate not in ws.merged_cells
                    )
                    ws.column_dimensions[column[0].column_letter].width = max(
                        width, 10) + 4
                except:
                    pass
